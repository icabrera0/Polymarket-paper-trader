"""
Ejecuta el backtesting del bot sobre mercados resueltos de Polymarket.

Uso:
    python scripts/run_backtest.py                    # 50 mercados, modo current
    python scripts/run_backtest.py --mode replay      # Noticias históricas (más lento)
    python scripts/run_backtest.py --markets 100      # Más mercados
    python scripts/run_backtest.py --balance 200      # Balance inicial personalizado
    python scripts/run_backtest.py --excel            # Exportar resultados a Excel

NOTAS IMPORTANTES:
  - El modo "current" tiene look-ahead bias (las noticias son de hoy, no del
    momento del mercado). Sirve para calibrar el pipeline, NO para medir
    la estrategia de forma real.
  - El modo "replay" es más realista pero depende de la cobertura histórica
    de GDELT, que es limitada para eventos nicho.
  - El backtest usa una DB separada (data/backtest.db) y no toca la producción.
  - Si usas Ollama, asegúrate de que el servicio esté corriendo.
  - Cada mercado analizado consume ~1 llamada al LLM. Con 50 mercados y
    Ollama: ~10-15 minutos. Con Anthropic: ~$0.05-0.10.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.backtester import Backtester  # noqa: E402
from src.config_loader import load_config  # noqa: E402


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Backtesting del Polymarket Paper Trading Bot"
    )
    parser.add_argument(
        "--mode",
        choices=["current", "replay"],
        default="current",
        help=(
            "current: noticias de hoy (look-ahead bias, rápido). "
            "replay: noticias históricas de GDELT (más realista, más lento)."
        ),
    )
    parser.add_argument(
        "--markets",
        type=int,
        default=50,
        help="Número de mercados resueltos a analizar (default: 50).",
    )
    parser.add_argument(
        "--balance",
        type=float,
        default=None,
        help="Balance inicial en EUR (default: el del config, 150€).",
    )
    parser.add_argument(
        "--excel",
        action="store_true",
        help="Exportar resultados a Excel en reports/backtest_YYYY-MM-DD.xlsx",
    )
    args = parser.parse_args()

    config = load_config()

    initial_balance = args.balance or config.paper_trading.initial_balance_eur

    print()
    print("═" * 65)
    print("  POLYMARKET PAPER TRADING BOT — BACKTESTING")
    print("═" * 65)
    print(f"  Modo:             {args.mode}")
    print(f"  Mercados:         {args.markets}")
    print(f"  Balance inicial:  €{initial_balance:.2f}")
    print(f"  LLM:              {config.llm.provider} ({config.llm.model})")
    print()

    if config.llm.provider == "ollama":
        from src.llm_client import OllamaClient, OllamaUnavailable
        try:
            OllamaClient(config).verify_setup()
            print("  ✓ Ollama OK")
        except OllamaUnavailable as exc:
            print(f"  ✗ Ollama no disponible: {exc}")
            print("    Ejecuta 'ollama serve' primero.")
            sys.exit(1)

    print(f"  Iniciando backtesting ({args.markets} mercados)...")
    print("  Esto puede tardar varios minutos. Ctrl+C para cancelar.")
    print()

    backtester = Backtester(
        config=config,
        mode=args.mode,
        max_markets=args.markets,
        initial_balance=initial_balance,
    )

    try:
        result = backtester.run()
    except KeyboardInterrupt:
        print("\n  Backtesting cancelado.")
        sys.exit(0)

    result.print_summary()

    if args.excel:
        _export_excel(result, config)


def _export_excel(result, config) -> None:
    """Exporta los resultados a un Excel básico de backtesting."""
    from datetime import datetime, timezone
    from pathlib import Path as P

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    output_dir = P(config.reports.output_directory)
    output_dir.mkdir(parents=True, exist_ok=True)
    fname = f"backtest_{datetime.now(timezone.utc).strftime('%Y-%m-%d_%H%M')}.xlsx"
    out_path = output_dir / fname

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen Backtest"

    # Cabecera
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    ws_summary["A1"] = "Parámetro"
    ws_summary["B1"] = "Valor"
    for cell in ["A1", "B1"]:
        ws_summary[cell].font = header_font
        ws_summary[cell].fill = header_fill

    kpis = [
        ("Modo", result.mode),
        ("Mercados analizados", result.markets_analyzed),
        ("Trades ejecutados", result.trades_executed),
        ("Win rate", f"{result.win_rate:.1%}"),
        ("P&L total", f"€{result.total_pnl_eur:+.2f}"),
        ("Balance inicial", f"€{result.initial_balance:.2f}"),
        ("Balance final", f"€{result.final_balance:.2f}"),
        ("Retorno total",
         f"{(result.final_balance - result.initial_balance) / result.initial_balance:+.2%}"),
        ("Max drawdown", f"{result.max_drawdown_pct:.2%}"),
        ("Sharpe ratio", f"{result.sharpe_ratio:.2f}"),
    ]
    for row, (k, v) in enumerate(kpis, 2):
        ws_summary[f"A{row}"] = k
        ws_summary[f"B{row}"] = v

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 18

    # Hoja de trades detallados
    ws_trades = wb.create_sheet("Trades Detallados")
    trade_headers = [
        "Mercado", "YES ganó", "Lado", "Entrada", "Salida",
        "Tamaño €", "P&L €", "P&L %",
        "Confianza", "Edge", "Artículos", "Low info", "Rec. LLM",
    ]
    for col_idx, header in enumerate(trade_headers, 1):
        cell = ws_trades.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    executed = [
        t for t in result.trades
        if t.decision == "OPEN_TRADE"
    ]
    for r_idx, t in enumerate(executed, 2):
        from openpyxl.utils import get_column_letter
        row_values = [
            t.market_question[:60],
            "Sí" if t.resolved_yes else "No",
            t.side.value if hasattr(t.side, "value") else str(t.side),
            t.entry_price_simulated,
            t.exit_price,
            t.size_eur,
            t.pnl_eur,
            t.pnl_pct,
            t.confidence,
            t.edge,
            t.num_articles,
            "Sí" if t.is_low_info else "No",
            t.llm_recommendation,
        ]
        for c_idx, value in enumerate(row_values, 1):
            ws_trades.cell(row=r_idx, column=c_idx, value=value)

        # Colores P&L
        pnl_cell = ws_trades.cell(row=r_idx, column=7)
        if (t.pnl_eur or 0) >= 0:
            pnl_cell.fill = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
        else:
            pnl_cell.fill = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")

    wb.save(out_path)
    print(f"  Excel exportado: {out_path}")


if __name__ == "__main__":
    main()
