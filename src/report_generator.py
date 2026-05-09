"""
Report Generator — informes Excel diarios del paper trading.

Genera un .xlsx con 5 hojas:

  1. Resumen Ejecutivo  — KPIs del día (balance, P&L, win rate, drawdown)
  2. Trades Detallados  — todas las posiciones cerradas hoy + abiertas vivas
  3. Análisis del LLM   — todos los MarketAnalysis del día con tokens consumidos
  4. Decisiones         — log completo (incluyendo NO_TRADE) para auditoría
  5. Evolución Balance  — curva de equity con gráfico

Decisiones de diseño:
- openpyxl porque necesitamos formato condicional + gráficos.
- Formulas Excel para totales (SUM, AVERAGE, COUNTIF) — NO hardcodear cálculos.
- Formato condicional verde/rojo en columnas P&L.
- Gráfico de línea para evolución de balance.
- Gráfico de barras para P&L por mercado.
- Filename: `YYYY-MM-DD_report.xlsx`.

El reporte se regenera por completo cada vez (no se actualiza incremental).
Para tiempo real usaremos el dashboard Streamlit en el módulo 11.
"""

from __future__ import annotations

from datetime import datetime, time, timedelta, timezone
from pathlib import Path
from typing import Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.config_loader import BotConfig
from src.database import Database


# =====================================================
# Estilos
# =====================================================

FONT_BASE = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="1F3864")

FILL_HEADER = PatternFill("solid", start_color="1F3864", end_color="1F3864")
FILL_KPI = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
FILL_GAIN = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
FILL_LOSS = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


# =====================================================
# ReportGenerator
# =====================================================


class ReportGenerator:
    """Genera el reporte Excel del paper trading."""

    def __init__(self, config: BotConfig, db: Database) -> None:
        self.config = config
        self.cfg = config.reports
        self.db = db
        self._log = logger.bind(module="report_generator")

        self.output_dir = Path(self.cfg.output_directory)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # =====================================================
    # Entry points
    # =====================================================

    def generate_daily_report(
        self,
        target_date: Optional[datetime] = None,
    ) -> Path:
        """Genera el reporte para `target_date` (o hoy UTC si None)."""
        if target_date is None:
            target_date = datetime.now(timezone.utc)

        # Rango: 00:00 UTC de target_date hasta 23:59:59 UTC
        day_start = datetime.combine(
            target_date.date(), time.min, tzinfo=timezone.utc
        )
        day_end = day_start + timedelta(days=1) - timedelta(microseconds=1)

        # Filename
        fname = target_date.strftime(self.cfg.filename_format)
        out_path = self.output_dir / fname

        # Datos de la DB
        all_trades = self.db.get_all_trades()
        balance_history = self.db.get_balance_history()

        # Trades del día (entrada O salida en el rango)
        trades_today = [
            t for t in all_trades
            if (
                (t.entry_timestamp >= day_start and t.entry_timestamp <= day_end)
                or (t.exit_timestamp and day_start <= t.exit_timestamp <= day_end)
            )
        ]

        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Borrar la hoja inicial

        self._sheet_executive_summary(wb, day_start, day_end, trades_today, balance_history)
        self._sheet_trades_detail(wb, trades_today, all_trades)
        self._sheet_llm_analyses(wb, day_start, day_end)
        self._sheet_decisions_log(wb, day_start, day_end)
        self._sheet_balance_evolution(wb, balance_history)

        wb.save(out_path)
        self._log.info("Reporte generado: {}", out_path)
        return out_path

    # =====================================================
    # Hoja 1: Resumen Ejecutivo
    # =====================================================

    def _sheet_executive_summary(
        self, wb: Workbook, day_start: datetime, day_end: datetime,
        trades_today: list, balance_history: list,
    ) -> None:
        ws = wb.create_sheet("Resumen Ejecutivo")

        # Título
        ws["A1"] = f"Reporte Paper Trading — {day_start.strftime('%Y-%m-%d')}"
        ws["A1"].font = FONT_TITLE
        ws.merge_cells("A1:D1")

        ws["A2"] = "Generado:"
        ws["B2"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        # Cabecera de KPIs
        row = 4
        ws[f"A{row}"] = "MÉTRICAS DEL DÍA"
        ws[f"A{row}"].font = FONT_HEADER
        ws[f"A{row}"].fill = FILL_HEADER
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        # Calcular KPIs (formulas no aplica aquí, los datos están en otras hojas)
        # Pero los valores finales SÍ los referenciamos para que sean dinámicos.

        # Balance inicial / final del día
        balance_start, balance_end = self._get_day_balance_bounds(
            balance_history, day_start, day_end,
        )
        peak_today = self._get_day_peak(balance_history, day_start, day_end)
        max_drawdown_today = self._get_day_max_drawdown(
            balance_history, day_start, day_end
        )

        closed_today = [t for t in trades_today if t.exit_timestamp]
        winners = [t for t in closed_today if (t.pnl_eur or 0) > 0]
        losers = [t for t in closed_today if (t.pnl_eur or 0) < 0]
        total_pnl_eur = sum((t.pnl_eur or 0) for t in closed_today)

        win_rate = len(winners) / len(closed_today) if closed_today else 0.0

        kpis = [
            ("Balance inicial del día", balance_start, "€"),
            ("Balance final del día", balance_end, "€"),
            ("P&L día", total_pnl_eur, "€"),
            ("P&L día %",
             (total_pnl_eur / balance_start) if balance_start > 0 else 0,
             "%"),
            ("Trades cerrados", len(closed_today), ""),
            ("Trades ganadores", len(winners), ""),
            ("Trades perdedores", len(losers), ""),
            ("Win rate", win_rate, "%"),
            ("Peak balance del día", peak_today, "€"),
            ("Drawdown máx del día", max_drawdown_today, "%"),
        ]

        for label, value, unit in kpis:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = FONT_BOLD
            ws[f"A{row}"].fill = FILL_KPI
            ws[f"A{row}"].alignment = LEFT

            ws[f"B{row}"] = value
            if unit == "€":
                ws[f"B{row}"].number_format = '€#,##0.00;[Red]-€#,##0.00'
            elif unit == "%":
                ws[f"B{row}"].number_format = "0.00%"
            else:
                ws[f"B{row}"].number_format = "#,##0"
            ws[f"B{row}"].alignment = RIGHT
            ws[f"B{row}"].font = FONT_BASE
            row += 1

        # Aplicar color condicional al P&L
        pnl_row = 4 + 1 + 3  # offset de cabecera + posición de "P&L día"
        ws[f"B{pnl_row}"].fill = (
            FILL_GAIN if total_pnl_eur >= 0 else FILL_LOSS
        )
        ws[f"B{pnl_row + 1}"].fill = (
            FILL_GAIN if total_pnl_eur >= 0 else FILL_LOSS
        )

        # Anchos
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12

    # =====================================================
    # Hoja 2: Trades Detallados
    # =====================================================

    def _sheet_trades_detail(
        self, wb: Workbook, trades_today: list, all_trades: list,
    ) -> None:
        ws = wb.create_sheet("Trades Detallados")

        headers = [
            "Trade ID", "Entrada", "Salida", "Mercado", "Token", "Lado",
            "Precio entrada", "Precio salida", "Tokens", "Tamaño €",
            "P&L €", "P&L %", "Duración (h)", "Motivo cierre",
            "Confianza", "Estado", "Notas",
        ]
        # Cabeceras
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = CENTER
            cell.border = BORDER_THIN

        # Datos
        for r_idx, t in enumerate(trades_today, 2):
            duration_h = ""
            if t.exit_timestamp and t.entry_timestamp:
                duration_h = (
                    t.exit_timestamp - t.entry_timestamp
                ).total_seconds() / 3600

            row_values = [
                t.trade_id[:8],
                t.entry_timestamp.strftime("%Y-%m-%d %H:%M") if t.entry_timestamp else "",
                t.exit_timestamp.strftime("%Y-%m-%d %H:%M") if t.exit_timestamp else "",
                t.market_question[:60],
                t.token_id[:10] + "..." if len(t.token_id) > 10 else t.token_id,
                t.side.value if t.side else "",
                t.entry_price,
                t.exit_price if t.exit_price is not None else "",
                round(t.tokens_quantity, 2),
                t.size_eur,
                t.pnl_eur if t.pnl_eur is not None else "",
                t.pnl_pct if t.pnl_pct is not None else "",
                round(duration_h, 2) if duration_h != "" else "",
                t.close_reason.value if t.close_reason else "",
                t.confidence,
                t.status.value if t.status else "",
                (t.exit_reason_text or t.entry_reason or "")[:80],
            ]
            for c_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = FONT_BASE
                cell.border = BORDER_THIN

            # Format de números
            ws.cell(row=r_idx, column=7).number_format = "0.0000"      # Precio entrada
            ws.cell(row=r_idx, column=8).number_format = "0.0000"      # Precio salida
            ws.cell(row=r_idx, column=10).number_format = "€#,##0.00"  # Tamaño
            ws.cell(row=r_idx, column=11).number_format = '€#,##0.00;[Red]-€#,##0.00'  # P&L €
            ws.cell(row=r_idx, column=12).number_format = "0.00%"      # P&L %

        # Tabla con filtros automáticos (si hay datos)
        if trades_today:
            n_rows = len(trades_today) + 1
            n_cols = len(headers)
            ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
            try:
                tbl = Table(displayName="TblTrades", ref=ref)
                tbl.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showRowStripes=True,
                )
                ws.add_table(tbl)
            except Exception as exc:
                self._log.debug("Tabla no aplicada (probablemente sin filas): {}", exc)

            # Formato condicional en columnas P&L (K=11, L=12)
            for col_letter, col_idx in [("K", 11), ("L", 12)]:
                rng = f"{col_letter}2:{col_letter}{n_rows}"
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(
                        operator="lessThan",
                        formula=["0"],
                        fill=FILL_LOSS,
                    ),
                )
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(
                        operator="greaterThan",
                        formula=["0"],
                        fill=FILL_GAIN,
                    ),
                )

        # Anchos
        widths = [12, 18, 18, 50, 16, 10, 14, 14, 10, 12, 12, 10, 14, 14, 10, 10, 50]
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # =====================================================
    # Hoja 3: Análisis del LLM
    # =====================================================

    def _sheet_llm_analyses(
        self, wb: Workbook, day_start: datetime, day_end: datetime,
    ) -> None:
        ws = wb.create_sheet("Análisis LLM")

        headers = [
            "Timestamp", "Mercado", "Precio YES", "Prob. consensus",
            "Edge", "Confianza", "Sentiment", "Impact",
            "Recomendación", "Timeframe", "Contradicciones",
            "Nº noticias", "Modelo LLM", "Tokens IN", "Tokens OUT",
            "Resumen",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = CENTER

        try:
            cur = self.db._conn.execute(
                """
                SELECT * FROM analyses_log
                WHERE timestamp >= ? AND timestamp <= ?
                ORDER BY timestamp ASC
                """,
                (day_start.isoformat(), day_end.isoformat()),
            )
            rows = cur.fetchall()
        except Exception as exc:
            self._log.warning("No se pudo leer analyses_log: {}", exc)
            rows = []

        for r_idx, row in enumerate(rows, 2):
            row_values = [
                row["timestamp"][:19],
                (row["market_question"] or "")[:60],
                row["current_yes_price"],
                row["consensus_probability_yes"],
                row["edge"],
                row["confidence"],
                row["sentiment_score"],
                row["impact_score"],
                row["recommendation"],
                row["timeframe"],
                "Sí" if row["contradictory_sources"] else "No",
                row["num_articles_analyzed"],
                row["llm_model"],
                row["llm_input_tokens"],
                row["llm_output_tokens"],
                (row["summary"] or "")[:120],
            ]
            for c_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = FONT_BASE
                cell.border = BORDER_THIN

            # Formatos
            ws.cell(row=r_idx, column=3).number_format = "0.000"
            ws.cell(row=r_idx, column=4).number_format = "0.000"
            ws.cell(row=r_idx, column=5).number_format = "0.000"
            ws.cell(row=r_idx, column=7).number_format = "0.00"
            ws.cell(row=r_idx, column=8).number_format = "0"

        # Color scale en confianza (col 6)
        if rows:
            n_rows = len(rows) + 1
            ws.conditional_formatting.add(
                f"F2:F{n_rows}",
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="F8696B",
                    mid_type="num", mid_value=50, mid_color="FFEB84",
                    end_type="num", end_value=100, end_color="63BE7B",
                ),
            )
            # Edge: verde positivo, rojo negativo
            ws.conditional_formatting.add(
                f"E2:E{n_rows}",
                CellIsRule(operator="lessThan", formula=["0"], fill=FILL_LOSS),
            )
            ws.conditional_formatting.add(
                f"E2:E{n_rows}",
                CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_GAIN),
            )

        widths = [20, 50, 12, 14, 10, 10, 10, 10, 18, 14, 14, 10, 22, 12, 12, 60]
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # =====================================================
    # Hoja 4: Decisiones (todas)
    # =====================================================

    def _sheet_decisions_log(
        self, wb: Workbook, day_start: datetime, day_end: datetime,
    ) -> None:
        ws = wb.create_sheet("Decisiones")

        headers = [
            "Timestamp", "Acción", "Mercado", "Lado", "Tamaño €",
            "Confianza", "Edge", "Skip reasons", "Justificación",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = CENTER

        try:
            cur = self.db._conn.execute(
                """
                SELECT * FROM decisions_log
                WHERE timestamp >= ? AND timestamp <= ?
                ORDER BY timestamp ASC
                """,
                (day_start.isoformat(), day_end.isoformat()),
            )
            rows = cur.fetchall()
        except Exception as exc:
            self._log.warning("No se pudo leer decisions_log: {}", exc)
            rows = []

        for r_idx, row in enumerate(rows, 2):
            row_values = [
                row["timestamp"][:19],
                row["action"],
                (row["market_question"] or "")[:60],
                row["side"] or "",
                row["size_eur"] or "",
                row["confidence"] or "",
                row["edge"] or 0,
                (row["skip_reasons"] or "[]")[:80],
                (row["rationale"] or "")[:200],
            ]
            for c_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = FONT_BASE
                cell.border = BORDER_THIN

            ws.cell(row=r_idx, column=5).number_format = "€#,##0.00"
            ws.cell(row=r_idx, column=7).number_format = "0.000"

            # Resaltar OPEN_TRADE en verde
            if row["action"] == "OPEN_TRADE":
                ws.cell(row=r_idx, column=2).fill = FILL_GAIN

        widths = [20, 14, 50, 10, 12, 12, 10, 40, 60]
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # =====================================================
    # Hoja 5: Evolución del Balance
    # =====================================================

    def _sheet_balance_evolution(self, wb: Workbook, balance_history: list) -> None:
        ws = wb.create_sheet("Evolución Balance")

        headers = ["Timestamp", "Balance €", "Peak €", "Drawdown %",
                   "Posiciones abiertas", "Evento"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = CENTER

        for r_idx, row in enumerate(balance_history, 2):
            row_values = [
                row["timestamp"][:19],
                row["balance_eur"],
                row["peak_balance"],
                row["drawdown_pct"],
                row["open_positions"],
                row["event"],
            ]
            for c_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = FONT_BASE
                cell.border = BORDER_THIN

            ws.cell(row=r_idx, column=2).number_format = "€#,##0.00"
            ws.cell(row=r_idx, column=3).number_format = "€#,##0.00"
            ws.cell(row=r_idx, column=4).number_format = "0.00%"

        # Color scale en drawdown
        if balance_history:
            n_rows = len(balance_history) + 1
            ws.conditional_formatting.add(
                f"D2:D{n_rows}",
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="63BE7B",
                    mid_type="num", mid_value=0.15, mid_color="FFEB84",
                    end_type="num", end_value=0.30, end_color="F8696B",
                ),
            )

            # Gráfico de línea: balance + peak
            chart = LineChart()
            chart.title = "Evolución del balance"
            chart.style = 12
            chart.y_axis.title = "EUR"
            chart.x_axis.title = "Tiempo"

            data = Reference(
                ws,
                min_col=2, max_col=3,
                min_row=1, max_row=n_rows,
            )
            cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 24
            chart.height = 12
            ws.add_chart(chart, f"H2")

        widths = [22, 14, 14, 14, 14, 18]
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # =====================================================
    # Helpers de KPI
    # =====================================================

    @staticmethod
    def _get_day_balance_bounds(
        history: list, day_start: datetime, day_end: datetime,
    ) -> tuple[float, float]:
        """Devuelve (balance_inicial, balance_final) del día."""
        if not history:
            return 0.0, 0.0
        # Último balance ANTES de day_start (o el primero si todo es del día)
        before = [h for h in history if h["timestamp"] < day_start.isoformat()]
        within = [
            h for h in history
            if day_start.isoformat() <= h["timestamp"] <= day_end.isoformat()
        ]
        if before:
            balance_start = float(before[-1]["balance_eur"])
        elif within:
            balance_start = float(within[0]["balance_eur"])
        else:
            balance_start = float(history[-1]["balance_eur"])
        balance_end = (
            float(within[-1]["balance_eur"]) if within else balance_start
        )
        return balance_start, balance_end

    @staticmethod
    def _get_day_peak(
        history: list, day_start: datetime, day_end: datetime,
    ) -> float:
        within = [
            h for h in history
            if day_start.isoformat() <= h["timestamp"] <= day_end.isoformat()
        ]
        if not within:
            return 0.0
        return max(float(h["peak_balance"]) for h in within)

    @staticmethod
    def _get_day_max_drawdown(
        history: list, day_start: datetime, day_end: datetime,
    ) -> float:
        within = [
            h for h in history
            if day_start.isoformat() <= h["timestamp"] <= day_end.isoformat()
        ]
        if not within:
            return 0.0
        return max(float(h["drawdown_pct"]) for h in within)
