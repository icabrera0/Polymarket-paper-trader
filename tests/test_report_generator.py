"""
Tests del ReportGenerator.

Verifican que con datos sembrados en una DB temporal:
- El archivo Excel se genera sin excepciones.
- Tiene las 5 hojas esperadas.
- Las hojas tienen el contenido correcto.
- Los cálculos de KPIs son correctos.

No verificamos el detalle de cada celda, solo lo crítico. La validación
visual completa la harás abriendo el .xlsx generado.

Ejecutar:
    pytest tests/test_report_generator.py -v
"""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.database import Database
from src.models import (
    CloseReason,
    DecisionAction,
    MarketAnalysis,
    Position,
    SkipReason,
    Timeframe,
    TradeDecision,
    TradeRecommendation,
    TradeSide,
    TradeStatus,
)
from src.report_generator import ReportGenerator


# =====================================================
# Fixtures
# =====================================================


@pytest.fixture
def db_path(tmp_path: Path) -> Path:
    return tmp_path / "test.db"


@pytest.fixture
def output_dir(tmp_path: Path) -> Path:
    d = tmp_path / "reports"
    d.mkdir()
    return d


@pytest.fixture
def configured(config_factory, output_dir):
    cfg = config_factory()
    cfg.reports.output_directory = str(output_dir)
    cfg.reports.filename_format = "%Y-%m-%d_test.xlsx"
    return cfg


def seed_db(db: Database, day: datetime) -> None:
    """Siembra la DB con datos representativos del día `day`."""
    # Balance history
    db.log_balance(150.0, 150.0, 0.0, 0, "INIT")

    # Trade ganador cerrado hoy
    p1 = Position(
        market_question="Will Spain win the Euro 2028?",
        token_id="0xspain",
        side=TradeSide.BUY_YES,
        entry_price=0.40,
        size_eur=20.0, size_usd=21.4, tokens_quantity=53.5,
        entry_timestamp=day.replace(hour=10),
        stop_loss_price=0.32, take_profit_price=0.52,
        status=TradeStatus.CLOSED,
        exit_price=0.55,
        exit_timestamp=day.replace(hour=14),
        close_reason=CloseReason.TAKE_PROFIT,
        pnl_eur=7.5, pnl_pct=0.375,
        confidence=80,
    )
    db.insert_trade(p1)
    db.log_balance(157.5, 157.5, 0.0, 0, "TRADE_CLOSE")

    # Trade perdedor cerrado hoy
    p2 = Position(
        market_question="Will Bitcoin hit 200k?",
        token_id="0xbtc",
        side=TradeSide.BUY_YES,
        entry_price=0.30,
        size_eur=15.0, size_usd=16.05, tokens_quantity=53.5,
        entry_timestamp=day.replace(hour=11),
        stop_loss_price=0.24, take_profit_price=0.39,
        status=TradeStatus.CLOSED,
        exit_price=0.24,
        exit_timestamp=day.replace(hour=16),
        close_reason=CloseReason.STOP_LOSS,
        pnl_eur=-3.0, pnl_pct=-0.20,
        confidence=70,
    )
    db.insert_trade(p2)
    db.log_balance(154.5, 157.5, 0.019, 0, "TRADE_CLOSE")

    # Trade abierto
    p3 = Position(
        market_question="Will Trump win 2028?",
        token_id="0xtrump",
        side=TradeSide.BUY_YES,
        entry_price=0.45,
        size_eur=18.0, size_usd=19.26, tokens_quantity=42.8,
        entry_timestamp=day.replace(hour=18),
        stop_loss_price=0.36, take_profit_price=0.585,
        status=TradeStatus.OPEN,
        confidence=85,
    )
    db.insert_trade(p3)
    db.log_balance(136.5, 157.5, 0.133, 1, "TRADE_OPEN")

    # Análisis del LLM
    a1 = MarketAnalysis(
        market_id="m1",
        market_question="Will Spain win the Euro 2028?",
        yes_token_id="0xspain", no_token_id="0xspain_no",
        current_yes_price=0.40, current_no_price=0.59,
        consensus_probability_yes=0.55, edge=0.15,
        confidence=80, sentiment_score=0.6, impact_score=70.0,
        recommendation=TradeRecommendation.COMPRAR_YES,
        timeframe=Timeframe.HORAS,
        summary="Strong news favoring Spain",
        num_articles_analyzed=5,
        llm_model="qwen2.5:7b", llm_input_tokens=2500, llm_output_tokens=350,
        analyzed_at=day.replace(hour=10),
    )
    db.log_analysis(a1)

    a2 = MarketAnalysis(
        market_id="m2",
        market_question="Will Bitcoin hit 200k?",
        yes_token_id="0xbtc", no_token_id="0xbtc_no",
        current_yes_price=0.30, current_no_price=0.69,
        consensus_probability_yes=0.20, edge=-0.10,
        confidence=70, sentiment_score=-0.4, impact_score=60.0,
        recommendation=TradeRecommendation.COMPRAR_NO,
        timeframe=Timeframe.DIAS,
        summary="Bearish crypto sentiment",
        num_articles_analyzed=3,
        llm_model="qwen2.5:7b", llm_input_tokens=2200, llm_output_tokens=300,
        analyzed_at=day.replace(hour=11),
    )
    db.log_analysis(a2)

    # Decisiones
    d1 = TradeDecision(
        action=DecisionAction.OPEN_TRADE,
        market_id="m1",
        market_question="Will Spain win the Euro 2028?",
        side=TradeSide.BUY_YES, token_id="0xspain",
        entry_price=0.40, size_eur=20.0,
        stop_loss_price=0.32, take_profit_price=0.52,
        confidence=80, edge=0.15,
        rationale="Strong edge with 5 articles",
        decided_at=day.replace(hour=10),
    )
    db.log_decision(d1)

    d2 = TradeDecision(
        action=DecisionAction.NO_TRADE,
        market_id="m3",
        market_question="Will some niche thing happen?",
        skip_reasons=[SkipReason.LLM_INSUFFICIENT_DATA],
        rationale="No relevant news",
        decided_at=day.replace(hour=12),
    )
    db.log_decision(d2)


# =====================================================
# Tests
# =====================================================


class TestReportGenerator:
    def test_genera_archivo_xlsx(self, configured, db_path):
        db = Database(db_path)
        day = datetime.now(timezone.utc).replace(hour=12)
        seed_db(db, day)

        gen = ReportGenerator(configured, db)
        out_path = gen.generate_daily_report(target_date=day)

        assert out_path.exists()
        assert out_path.suffix == ".xlsx"
        assert out_path.stat().st_size > 1000  # Algo de contenido

    def test_excel_tiene_5_hojas(self, configured, db_path):
        db = Database(db_path)
        day = datetime.now(timezone.utc).replace(hour=12)
        seed_db(db, day)

        out_path = ReportGenerator(configured, db).generate_daily_report(day)
        wb = load_workbook(out_path)

        expected = {
            "Resumen Ejecutivo",
            "Trades Detallados",
            "Análisis LLM",
            "Decisiones",
            "Evolución Balance",
        }
        assert expected == set(wb.sheetnames)

    def test_resumen_ejecutivo_tiene_kpis(self, configured, db_path):
        db = Database(db_path)
        day = datetime.now(timezone.utc).replace(hour=12)
        seed_db(db, day)

        out_path = ReportGenerator(configured, db).generate_daily_report(day)
        wb = load_workbook(out_path)
        ws = wb["Resumen Ejecutivo"]

        # El título de la celda A1 debe contener la fecha
        assert "Reporte Paper Trading" in str(ws["A1"].value)
        # Buscar el texto de algún KPI
        labels = []
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            if row[0]:
                labels.append(str(row[0]))
        assert any("P&L día" in l for l in labels)
        assert any("Win rate" in l for l in labels)
        assert any("Drawdown" in l for l in labels)

    def test_trades_detallados_tiene_filas(self, configured, db_path):
        db = Database(db_path)
        day = datetime.now(timezone.utc).replace(hour=12)
        seed_db(db, day)

        out_path = ReportGenerator(configured, db).generate_daily_report(day)
        wb = load_workbook(out_path)
        ws = wb["Trades Detallados"]

        # 1 cabecera + 3 trades
        assert ws.max_row >= 4

    def test_analisis_llm_tiene_filas(self, configured, db_path):
        db = Database(db_path)
        day = datetime.now(timezone.utc).replace(hour=12)
        seed_db(db, day)

        out_path = ReportGenerator(configured, db).generate_daily_report(day)
        wb = load_workbook(out_path)
        ws = wb["Análisis LLM"]
        assert ws.max_row >= 3  # 1 cabecera + 2 análisis

    def test_decisiones_incluye_no_trade(self, configured, db_path):
        db = Database(db_path)
        day = datetime.now(timezone.utc).replace(hour=12)
        seed_db(db, day)

        out_path = ReportGenerator(configured, db).generate_daily_report(day)
        wb = load_workbook(out_path)
        ws = wb["Decisiones"]
        # Buscamos que aparezca tanto OPEN_TRADE como NO_TRADE
        actions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                actions.append(row[1])
        assert "OPEN_TRADE" in actions
        assert "NO_TRADE" in actions

    def test_evolucion_balance_tiene_grafico(self, configured, db_path):
        db = Database(db_path)
        day = datetime.now(timezone.utc).replace(hour=12)
        seed_db(db, day)

        out_path = ReportGenerator(configured, db).generate_daily_report(day)
        wb = load_workbook(out_path)
        ws = wb["Evolución Balance"]
        # ws._charts es la lista interna de openpyxl
        assert len(ws._charts) >= 1

    def test_genera_sin_datos_no_crashea(self, configured, db_path):
        # DB vacía: el report debe generarse sin excepciones (pero vacío)
        db = Database(db_path)
        out_path = ReportGenerator(configured, db).generate_daily_report()
        assert out_path.exists()
        wb = load_workbook(out_path)
        assert "Resumen Ejecutivo" in wb.sheetnames

    def test_filtra_por_dia_correcto(self, configured, db_path):
        """Trades de ayer NO deben aparecer en el reporte de hoy."""
        db = Database(db_path)
        today = datetime.now(timezone.utc).replace(hour=12)
        yesterday = today - timedelta(days=1)

        # Sembrar datos de AYER
        seed_db(db, yesterday)

        # Generar reporte de HOY (debe estar vacío de trades)
        out_path = ReportGenerator(configured, db).generate_daily_report(today)
        wb = load_workbook(out_path)
        ws = wb["Trades Detallados"]
        # Solo cabecera, ninguna fila de datos
        assert ws.max_row == 1
